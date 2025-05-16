import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import operator
import json
import os
import sys
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import threading
from concurrent.futures import ThreadPoolExecutor
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np
from datetime import datetime
from queue import Queue


# Modern color palette
BG_COLOR = "#f8f9fa"
PRIMARY_COLOR = "#6c757d"
SECONDARY_COLOR = "#343a40"
ACCENT_COLOR = "#0d6efd"
SUCCESS_COLOR = "#198754"
WARNING_COLOR = "#ffc107"
DANGER_COLOR = "#dc3545"
TEXT_COLOR = "#212529"
FONT_NAME = "Segoe UI"
FONT_SIZE = 10


def resource_path(relative_path):
    """Get absolute path to resources for PyInstaller"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


RULES_FILE = resource_path("djezzy_rules.json")

# ====== RULES MANAGEMENT ======
DEFAULT_RULES = {
    "2G": [
        {"kpi": "2G_CSSR_CS(%)", "operator": ">",
         "threshold": 98, "count_threshold": 100},
        {"kpi": "CDR_OPTIMUM", "operator": "<",
            "threshold": 1, "count_threshold": 50},
        {"kpi": "HSR_OPTIMUM", "operator": ">=",
            "threshold": 98, "count_threshold": 100},
    ],
    "3G": [
        {"kpi": "Call Setup Success Rate PS_OPTIMUM", "operator": ">", "threshold": 98,
         "count_column": "PS_Attempts", "count_threshold": 100},
        {"kpi": "Call Setup Success Rate CS_OPTIMUM", "operator": ">", "threshold": 98,
         "count_column": "CS_Attempts", "count_threshold": 100},
        {"kpi": "Call Drop Rate CS_OPTIMUM", "operator": "<", "threshold": 1,
         "count_column": "Dropped_Calls", "count_threshold": 50},
        {"kpi": "RTWP_optimum(dBm)", "operator": "<",
         "threshold": -95, "count_threshold": 0},
        {"kpi": "EVQI Bad+Poor_Optimum (%)", "operator": "<",
         "threshold": 2, "count_threshold": 100},
    ],
    "4G": [
        {"kpi": "LTE Setup Success Rate_OPTIMUM(%)", "operator": ">", "threshold": 99,
         "count_column": "LTE_Attempts", "count_threshold": 100},
        {"kpi": "LTE Call Drop Rate_OPTIMUM", "operator": "<", "threshold": 0.8,
         "count_column": "LTE_Drops", "count_threshold": 50},
        {"kpi": "CSFB Success Rate_OPTIMUM(%)", "operator": ">",
         "threshold": 99.5, "count_threshold": 100},
    ]
}

# Load or initialize rules
try:
    with open(RULES_FILE, 'r') as f:
        rules_map = json.load(f)
except (FileNotFoundError, json.JSONDecodeError):
    rules_map = DEFAULT_RULES
    with open(RULES_FILE, 'w') as f:
        json.dump(rules_map, f, indent=4)

ops = {
    ">=": operator.ge, "<=": operator.le,
    ">": operator.gt, "<": operator.lt, "==": operator.eq
}

# ====== App Configuration ======


class AppConfig:
    APP_NAME = "Cell Performance Analyzer "
    COMPANY = "Access Network West Optimization Team"
    VERSION = "1.0"

    # Modern color palette
    COLORS = {
        'bg': "#f8f9fa",
        'primary': "#6c757d",
        'secondary': "#343a40",
        'accent': "#0d6efd",
        'success': "#198754",
        'warning': "#ffc107",
        'danger': "#dc3545",
        'text': "#212529",
        'card_bg': "#ffffff",
        'nav_bg': "#2c3e50"
    }

    FONT = {
        'name': "Segoe UI",
        'size': 10,
        'title_size': 12
    }

    # Threshold values
    THRESHOLDS = {
        '2G': {
            'CSSR_CS': 98,
            'CDR': 1,
            'HSR': 98
        },
        '3G': {
            'CSSR_PS': 98,
            'CSSR_CS': 98,
            'CDR_CS': 1,
            'RTWP': -95,
            'EVQI': 2
        },
        '4G': {
            'LTE_SSR': 99,
            'LTE_CDR': 0.8,
            'CSFB_SR': 99.5
        }
    }

    # File paths
    @staticmethod
    def resource_path(relative_path):
        """Get absolute path to resources for PyInstaller"""
        if hasattr(sys, '_MEIPASS'):
            return os.path.join(sys._MEIPASS, relative_path)
        return os.path.join(os.path.abspath("."), relative_path)

# ====== Data Analysis Engine ======


class CellAnalyzer:
    def __init__(self):
        self.ops = {
            ">=": operator.ge, "<=": operator.le,
            ">": operator.gt, "<": operator.lt, "==": operator.eq
        }
        self.load_rules()
        self.data_cache = {}
        self.analysis_cache = {}

    def load_rules(self):
        """Load analysis rules from JSON file"""
        self.rules_file = AppConfig.resource_path("djezzy_rules.json")
        try:
            with open(self.rules_file, 'r') as f:
                self.rules = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            self.rules = self.get_default_rules()
            self.save_rules()

    def get_default_rules(self):
        """Return default analysis rules"""
        return {
            "2G": [
                {"kpi": "2G_CSSR_CS(%)", "operator": ">",
                 "threshold": 98, "count_threshold": 100},
                {"kpi": "CDR_OPTIMUM", "operator": "<",
                    "threshold": 1, "count_threshold": 50},
                {"kpi": "HSR_OPTIMUM", "operator": ">=",
                    "threshold": 98, "count_threshold": 100},
            ],
            "3G": [
                {"kpi": "Call Setup Success Rate PS_OPTIMUM", "operator": ">", "threshold": 98,
                 "count_column": "PS_Attempts", "count_threshold": 100},
                {"kpi": "Call Setup Success Rate CS_OPTIMUM", "operator": ">", "threshold": 98,
                 "count_column": "CS_Attempts", "count_threshold": 100},
                {"kpi": "Call Drop Rate CS_OPTIMUM", "operator": "<", "threshold": 1,
                 "count_column": "Dropped_Calls", "count_threshold": 50},
                {"kpi": "RTWP_optimum(dBm)", "operator": "<",
                 "threshold": -95, "count_threshold": 0},
                {"kpi": "EVQI Bad+Poor_Optimum (%)", "operator": "<",
                 "threshold": 2, "count_threshold": 100},
            ],
            "4G": [
                {"kpi": "LTE Setup Success Rate_OPTIMUM(%)", "operator": ">", "threshold": 99,
                 "count_column": "LTE_Attempts", "count_threshold": 100},
                {"kpi": "LTE Call Drop Rate_OPTIMUM", "operator": "<", "threshold": 0.8,
                 "count_column": "LTE_Drops", "count_threshold": 50},
                {"kpi": "CSFB Success Rate_OPTIMUM(%)", "operator": ">",
                 "threshold": 99.5, "count_threshold": 100},
            ]
        }

    def save_rules(self):
        """Save rules to JSON file"""
        with open(self.rules_file, 'w') as f:
            json.dump(self.rules, f, indent=4)

    def analyze_technology(self, file_path, tech):
        """Analyze a single technology and return summary and details"""
        try:
            # Check cache first
            cache_key = f"{file_path}_{tech}"
            if cache_key in self.analysis_cache:
                return self.analysis_cache[cache_key]

            # Load data with caching
            if cache_key not in self.data_cache:
                xls = pd.ExcelFile(file_path)
                # sheet name must be dynamic
                df = pd.read_excel(xls)
                df["Date"] = pd.to_datetime(df["Date"], dayfirst=True)
                self.data_cache[cache_key] = df
            else:
                df = self.data_cache[cache_key]

            cell_names = df["Cell Name"].unique()
            total_cells = len(cell_names)

            summary = {
                "technology": tech,
                "total_cells": total_cells,
                "critical": 0,
                "warning": 0,
                "healthy": total_cells,
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }

            cell_details = []
            problematic_cells = set()

            # Parallel processing of rules with progress tracking
            with ThreadPoolExecutor(max_workers=4) as executor:
                futures = []
                for rule in self.rules.get(tech, []):
                    futures.append(executor.submit(self.analyze_kpi, df, rule))

                for future in futures:
                    res = future.result()
                    if res is not None and not res.empty:
                        for _, row in res.iterrows():
                            cell_name = row["Cell Name"]
                            problematic_cells.add(cell_name)

                            if row["Status"] == "Critical":
                                summary["critical"] += 1
                            else:
                                summary["warning"] += 1

                            cell_details.append(row.to_dict())

            # Update healthy count
            summary["healthy"] = summary["total_cells"] - \
                len(problematic_cells)

            # Cache results
            result = (summary, cell_details)
            self.analysis_cache[cache_key] = result
            return result

        except Exception as e:
            print(f"Error processing {tech} sheet: {str(e)}")
            return None, None

    def analyze_kpi(self, df, rule):
        """Parallel processing of KPI analysis"""
        try:
            latest_dates = sorted(df["Date"].unique())[-7:]  # Last 7 days
            cell_names = df["Cell Name"].unique()

            with ThreadPoolExecutor(max_workers=4) as executor:
                futures = [executor.submit(self.process_cell_data, cell, df, rule, latest_dates)
                           for cell in cell_names]
                results = [future.result()
                           for future in futures if future.result() is not None]

            if not results:
                return None

            result_df = pd.DataFrame(results)
            return result_df.sort_values(by=["Score"], ascending=False)
        except Exception as e:
            print(f"Error in analyze_kpi: {str(e)}")
            return None

    def process_cell_data(self, cell_name, df, rule, latest_dates):
        """Process data for a single cell"""
        cell_data = df[df["Cell Name"] == cell_name]
        bad_days = 0
        bad_number = 0
        daily_values = {}
        last_5_bad = 0
        last_day = False

        for i, day in enumerate(latest_dates, 1):
            day_data = cell_data[cell_data["Date"] == day]
            col_name = f"d{i}"

            if not day_data.empty:
                value = day_data.iloc[0].get(rule["kpi"], None)
                op_func = self.ops[rule["operator"]]

                if pd.notna(value):
                    is_bad = not op_func(value, rule["threshold"])
                    daily_values[col_name] = value

                    if "count_column" in rule:
                        count_val = day_data.iloc[0].get(
                            rule["count_column"], 0)
                        daily_values[f"{col_name}_count"] = count_val
                    else:
                        daily_values[f"{col_name}_count"] = "-"

                    if is_bad:
                        bad_days += 1
                        if i > 2:  # Last 5 days
                            last_5_bad += 1
                        if "count_column" in rule and count_val > rule["count_threshold"]:
                            bad_number += 1
                        if i == 7:  # Last day
                            last_day = True
                else:
                    daily_values[col_name] = "No Data"
                    daily_values[f"{col_name}_count"] = ""
            else:
                daily_values[col_name] = "No Data"
                daily_values[f"{col_name}_count"] = ""

        if last_5_bad >= 3 and last_day:
            return {
                "Cell Name": cell_name,
                "KPI": rule["kpi"],
                "Bad_days": bad_days,
                "failure_number": bad_number,
                "Last_5_days": last_5_bad,
                "Score": bad_days + last_5_bad + bad_number,
                **daily_values,
                "Status": "Critical" if last_5_bad == 5 else "Warning"
            }
        return None

    def get_worst_cells_for_kpi(self, cell_details, kpi, n=5):
        """Get the worst performing cells for a specific KPI"""
        kpi_cells = [cell for cell in cell_details if cell["KPI"] == kpi]
        sorted_cells = sorted(
            kpi_cells, key=lambda x: x["Score"], reverse=True)
        return sorted_cells[:n]

# ====== GUI Components ======


class ModernButton(ttk.Button):
    """Styled button component"""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.configure(style="Accent.TButton")


class Card(ttk.Frame):
    """Card component for dashboard"""

    def __init__(self, parent, title, value=None, color=None, **kwargs):
        super().__init__(parent, style="Card.TFrame", padding=15, **kwargs)

        # Title label
        ttk.Label(self, text=title, style="CardTitle.TLabel").pack(anchor="w")

        # Value label
        self.value_var = tk.StringVar(value=value if value else "")
        value_label = ttk.Label(self, textvariable=self.value_var,
                                style="CardValue.TLabel",
                                foreground=color if color else AppConfig.COLORS['text'])
        value_label.pack(anchor="w", pady=(10, 0))

    def update_value(self, new_value):
        self.value_var.set(new_value)


class PerformanceGraph(ttk.Frame):
    """Custom graph component for performance visualization"""

    def __init__(self, parent, title, width=5, height=3, dpi=100):
        super().__init__(parent, style="Card.TFrame", padding=15)
        self.title = title
        self.width = width
        self.height = height
        self.dpi = dpi

        ttk.Label(self, text=title, style="CardTitle.TLabel").pack(anchor="w")

        self.fig, self.ax = plt.subplots(figsize=(width, height), dpi=dpi)
        self.fig.patch.set_facecolor('white')
        self.canvas = FigureCanvasTkAgg(self.fig, master=self)
        self.canvas.get_tk_widget().pack(fill="both", expand=True, pady=10)

    def update_chart(self, data, x_labels, title=None, ylabel=None):
        """Update the chart with new data"""
        self.ax.clear()

        if isinstance(data, dict):
            # Single dataset
            self.ax.plot(x_labels, list(data.values()), marker='o')
        else:
            # Multiple datasets
            for cell_data in data:
                self.ax.plot(
                    x_labels, cell_data['values'], marker='o', label=cell_data['label'])
            self.ax.legend()

        if title:
            self.ax.set_title(title)
        if ylabel:
            self.ax.set_ylabel(ylabel)

        self.ax.grid(True, linestyle='--', alpha=0.6)
        self.fig.tight_layout()
        self.canvas.draw()


class CellPerformanceApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(
            f"{AppConfig.APP_NAME} | {AppConfig.COMPANY} | v{AppConfig.VERSION}")
        self.geometry("1600x900")
        self.minsize(1200, 700)
        self.configure(bg=AppConfig.COLORS['bg'])

        # Initialize analyzer
        self.analyzer = CellAnalyzer()

        # Data storage
        self.current_file = ""
        self.selected_tech = ""
        self.summary_data = None
        self.cell_details = None
        self.analysis_results = None
        self.status_queue = Queue()
        self.dashboard_ready = False

        # Setup UI
        self._setup_styles()
        self._create_widgets()
        self.show_analysis()  # Start with analysis tab

        # Start status update thread
        self.after(100, self._process_status_queue)

    def _setup_styles(self):
        """Configure modern UI styles"""
        self.style = ttk.Style()
        self.style.theme_use("clam")

        # Base styles
        self.style.configure(".",
                             font=(AppConfig.FONT['name'], AppConfig.FONT['size']))
        self.style.configure("TFrame",
                             background=AppConfig.COLORS['bg'])
        self.style.configure("TLabel",
                             background=AppConfig.COLORS['bg'],
                             foreground=AppConfig.COLORS['text'])
        self.style.configure("TButton",
                             padding=6)

        # Accent button
        self.style.configure("Accent.TButton",
                             background=AppConfig.COLORS['accent'],
                             foreground="white",
                             font=(AppConfig.FONT['name'], AppConfig.FONT['size'], "bold"))
        self.style.map("Accent.TButton",
                       background=[("active", "#0b5ed7"), ("pressed", "#0a58ca")])

        # Card styles
        self.style.configure("Card.TFrame",
                             background=AppConfig.COLORS['card_bg'],
                             borderwidth=1,
                             relief="solid")
        self.style.configure("CardTitle.TLabel",
                             background=AppConfig.COLORS['card_bg'],
                             foreground=AppConfig.COLORS['secondary'],
                             font=(AppConfig.FONT['name'], AppConfig.FONT['size'], "bold"))
        self.style.configure("CardValue.TLabel",
                             background=AppConfig.COLORS['card_bg'],
                             foreground=AppConfig.COLORS['primary'],
                             font=(AppConfig.FONT['name'], 14, "bold"))

        # Navigation styles
        self.style.configure("Nav.TFrame",
                             background=AppConfig.COLORS['nav_bg'])
        self.style.configure("Nav.TLabel",
                             background=AppConfig.COLORS['nav_bg'],
                             foreground="white")
        self.style.configure("Nav.TButton",
                             background=AppConfig.COLORS['nav_bg'],
                             foreground="white",
                             font=(AppConfig.FONT['name'], AppConfig.FONT['size']))
        self.style.map("Nav.TButton",
                       background=[("active", AppConfig.COLORS['primary']),
                                   ("pressed", AppConfig.COLORS['primary'])])

    def _create_widgets(self):
        """Build the interface"""
        # Header
        header_frame = ttk.Frame(self, padding=(20, 10), style="Nav.TFrame")
        header_frame.pack(fill="x", side="top")

        # App logo and title
        logo_frame = ttk.Frame(header_frame, style="Nav.TFrame")
        logo_frame.pack(side="left")

        logo_label = ttk.Label(logo_frame, text="CPA",
                               style="Nav.TLabel",
                               font=(AppConfig.FONT['name'], 20, "bold"))
        logo_label.pack(side="left", padx=10)

        title_label = ttk.Label(logo_frame, text=AppConfig.APP_NAME,
                                style="Nav.TLabel",
                                font=(AppConfig.FONT['name'], 16))
        title_label.pack(side="left", padx=5)

        # Version label
        version_label = ttk.Label(header_frame, text=f"v{AppConfig.VERSION}",
                                  style="Nav.TLabel")
        version_label.pack(side="right", padx=20)

        # Navigation buttons
        nav_frame = ttk.Frame(header_frame, style="Nav.TFrame")
        nav_frame.pack(side="right")

        self.dashboard_btn = ttk.Button(nav_frame, text="Dashboard",
                                        style="Nav.TButton",
                                        command=self.show_dashboard)
        self.dashboard_btn.pack(side="left", padx=5)

        self.analysis_btn = ttk.Button(nav_frame, text="Analysis",
                                       style="Nav.TButton",
                                       command=self.show_analysis)
        self.analysis_btn.pack(side="left", padx=5)

        self.cell_analysis_btn = ttk.Button(nav_frame, text="Cell Analysis",
                                            style="Nav.TButton",
                                            command=self.show_cell_analysis)
        self.cell_analysis_btn.pack(side="left", padx=5)

        self.rules_btn = ttk.Button(nav_frame, text="Rules",
                                    style="Nav.TButton",
                                    command=self._open_rule_editor)
        self.rules_btn.pack(side="left", padx=5)

        # Main content area
        self.main_frame = ttk.Frame(self)
        self.main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # Create frames (initially hidden)
        self.create_analysis_frame()
        self.create_dashboard_frame()
        self.create_cell_analysis_frame()

    def create_analysis_frame(self):
        """Create the analysis frame"""
        self.analysis_frame = ttk.Frame(self.main_frame)

        # File selection card
        file_card = ttk.Frame(self.analysis_frame,
                              style="Card.TFrame", padding=15)
        file_card.pack(fill="x", pady=(0, 15))

        ttk.Label(file_card, text="Select Data File",
                  style="CardTitle.TLabel").pack(anchor="w")

        self.file_var = tk.StringVar()
        file_entry_frame = ttk.Frame(file_card)
        file_entry_frame.pack(fill="x", pady=10)

        file_entry = ttk.Entry(file_entry_frame, textvariable=self.file_var)
        file_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

        browse_btn = ModernButton(
            file_entry_frame, text="Browse", command=self._browse_file)
        browse_btn.pack(side="right")

        # Technology selection
        tech_frame = ttk.Frame(self.analysis_frame,
                               style="Card.TFrame", padding=15)
        tech_frame.pack(fill="x", pady=(0, 20))

        ttk.Label(tech_frame, text="Select Technology:",
                  style="CardTitle.TLabel").pack(anchor="w")
        self.tech_var = tk.StringVar(value="2G")
        tech_combo = ttk.Combobox(tech_frame, textvariable=self.tech_var,
                                  values=["2G", "3G", "4G"], state="readonly")
        tech_combo.pack(anchor="w", pady=5, fill="x")

        # Action buttons
        button_frame = ttk.Frame(self.analysis_frame)
        button_frame.pack(fill="x", pady=(10, 0))

        ModernButton(button_frame, text="Start Analysis",
                     command=self._start_analysis).pack(side="left", padx=5)

        self.progress_bar = ttk.Progressbar(
            button_frame, mode="indeterminate", length=200)
        self.progress_bar.pack(side="left", expand=True, padx=10)

        ModernButton(button_frame, text="Manage Rules",
                     command=self._open_rule_editor).pack(side="right", padx=5)

    def create_dashboard_frame(self):
        """Create the professional dashboard"""
        self.dashboard_frame = ttk.Frame(self.main_frame)

        # Summary section
        summary_frame = ttk.Frame(self.dashboard_frame)
        summary_frame.pack(fill="x", pady=(0, 20))

        # Summary cards
        self.summary_cards = {
            'total': Card(summary_frame, "Total Cells", "0"),
            'healthy': Card(summary_frame, "Healthy Cells", "0", AppConfig.COLORS['success']),
            'warning': Card(summary_frame, "Warning Cells", "0", AppConfig.COLORS['warning']),
            'critical': Card(summary_frame, "Critical Cells", "0", AppConfig.COLORS['danger'])
        }

        for i, card in enumerate(self.summary_cards.values()):
            card.grid(row=0, column=i, padx=5, sticky="nsew")
            summary_frame.grid_columnconfigure(i, weight=1)

        # Charts section
        charts_frame = ttk.Frame(self.dashboard_frame)
        charts_frame.pack(fill="both", expand=True, pady=(0, 20))

        # Health distribution chart
        self.health_chart = PerformanceGraph(
            charts_frame, "Health Distribution")
        self.health_chart.pack(side="left", fill="both", expand=True, padx=5)

        # Status over time chart with KPI selector
        self.status_chart_frame = ttk.Frame(
            charts_frame, style="Card.TFrame", padding=15)
        self.status_chart_frame.pack(
            side="left", fill="both", expand=True, padx=5)

        # Chart title and controls
        chart_header = ttk.Frame(self.status_chart_frame)
        chart_header.pack(fill="x", pady=(0, 10))

        ttk.Label(chart_header, text="Status Over Time",
                  style="CardTitle.TLabel").pack(side="left")

        # KPI selector
        self.kpi_chart_var = tk.StringVar()
        self.kpi_chart_combo = ttk.Combobox(
            chart_header, textvariable=self.kpi_chart_var, state="readonly")
        self.kpi_chart_combo.pack(side="right", padx=10)
        self.kpi_chart_combo.bind(
            "<<ComboboxSelected>>", self._update_status_chart)

        # Number of cells to show
        self.cell_count_var = tk.IntVar(value=5)
        cell_count_frame = ttk.Frame(chart_header)
        cell_count_frame.pack(side="right", padx=10)
        ttk.Label(cell_count_frame, text="Show top:").pack(side="left")
        ttk.Spinbox(cell_count_frame, from_=1, to=10, width=3, textvariable=self.cell_count_var,
                    command=self._update_status_chart).pack(side="left")

        # The actual chart
        self.status_chart = PerformanceGraph(
            self.status_chart_frame, "", width=6, height=4)
        self.status_chart.pack(fill="both", expand=True)

        # Action buttons frame
        action_frame = ttk.Frame(self.dashboard_frame)
        action_frame.pack(fill="x", pady=(10, 0))

        ModernButton(action_frame, text="Export Full Report",
                     command=self._export_full_report).pack(side="right", padx=5)

    def create_cell_analysis_frame(self):
        """Create the enhanced cell analysis frame"""
        self.cell_analysis_frame = ttk.Frame(self.main_frame)

        # Filter controls
        filter_frame = ttk.Frame(
            self.cell_analysis_frame, style="Card.TFrame", padding=15)
        filter_frame.pack(fill="x", pady=(0, 15))

        ttk.Label(filter_frame, text="Filter Cells:",
                  style="CardTitle.TLabel").pack(anchor="w")

        # Status filter
        status_frame = ttk.Frame(filter_frame)
        status_frame.pack(fill="x", pady=5)
        ttk.Label(status_frame, text="Status:").pack(side="left", padx=(0, 10))

        self.status_var = tk.StringVar(value="All")
        ttk.Radiobutton(status_frame, text="All", variable=self.status_var, value="All",
                        command=self._filter_cells).pack(side="left", padx=5)
        ttk.Radiobutton(status_frame, text="Critical", variable=self.status_var, value="Critical",
                        command=self._filter_cells).pack(side="left", padx=5)
        ttk.Radiobutton(status_frame, text="Warning", variable=self.status_var, value="Warning",
                        command=self._filter_cells).pack(side="left", padx=5)

        # KPI filter
        kpi_frame = ttk.Frame(filter_frame)
        kpi_frame.pack(fill="x", pady=5)
        ttk.Label(kpi_frame, text="KPI:").pack(side="left", padx=(0, 10))

        self.kpi_var = tk.StringVar()
        self.kpi_combo = ttk.Combobox(
            kpi_frame, textvariable=self.kpi_var, state="readonly")
        self.kpi_combo.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.kpi_combo.bind("<<ComboboxSelected>>",
                            lambda e: self._filter_cells())

        # Cell details table
        table_frame = ttk.Frame(self.cell_analysis_frame)
        table_frame.pack(fill="both", expand=True)

        # Create Treeview with scrollbars
        columns = [
            ("cell", "Cell Name", 150),
            ("kpi", "KPI", 200),
            ("status", "Status", 80),
            ("score", "Score", 60),
            ("bad_days", "Bad Days", 80),
            ("d1", "D1", 60), ("d1_count", "Count", 60),
            ("d2", "D2", 60), ("d2_count", "Count", 60),
            ("d3", "D3", 60), ("d3_count", "Count", 60),
            ("d4", "D4", 60), ("d4_count", "Count", 60),
            ("d5", "D5", 60), ("d5_count", "Count", 60),
            ("d6", "D6", 60), ("d6_count", "Count", 60),
            ("d7", "D7", 60), ("d7_count", "Count", 60)
        ]

        self.cell_tree = ttk.Treeview(
            table_frame,
            columns=[col[0] for col in columns],
            show="headings",
            selectmode="extended",
            height=20
        )

        # Configure columns
        for col_id, heading, width in columns:
            self.cell_tree.heading(col_id, text=heading)
            self.cell_tree.column(col_id, width=width, anchor="center")

        # Add scrollbars
        yscroll = ttk.Scrollbar(
            table_frame, orient="vertical", command=self.cell_tree.yview)
        xscroll = ttk.Scrollbar(
            table_frame, orient="horizontal", command=self.cell_tree.xview)
        self.cell_tree.configure(yscroll=yscroll.set, xscroll=xscroll.set)

        self.cell_tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")

        # Configure grid weights
        table_frame.grid_columnconfigure(0, weight=1)
        table_frame.grid_rowconfigure(0, weight=1)

        # Action buttons
        action_frame = ttk.Frame(self.cell_analysis_frame)
        action_frame.pack(fill="x", pady=(10, 0))

        ModernButton(action_frame, text="Export Selected Cells",
                     command=self._export_selected_cells).pack(side="right", padx=5)

    # ====== Navigation Methods ======
    def show_analysis(self):
        """Show the analysis view"""
        self.dashboard_frame.pack_forget()
        self.cell_analysis_frame.pack_forget()
        self.analysis_frame.pack(fill="both", expand=True)
        self._update_button_states("analysis")

    def show_dashboard(self):
        """Show the dashboard view with optimized loading"""
        # Hide other frames first for immediate UI response
        self.analysis_frame.pack_forget()
        self.cell_analysis_frame.pack_forget()
        self.dashboard_frame.pack(fill="both", expand=True)
        self._update_button_states("dashboard")

        # Show loading state immediately
        for card in self.summary_cards.values():
            card.update_value("Loading...")

        self.health_chart.ax.clear()
        self.health_chart.ax.text(0.5, 0.5, "Loading data...",
                                  ha='center', va='center')
        self.health_chart.canvas.draw()

        self.status_chart.ax.clear()
        self.status_chart.ax.text(0.5, 0.5, "Loading data...",
                                  ha='center', va='center')
        self.status_chart.canvas.draw()

        # Update UI first, then load data in background
        self.update_idletasks()
        self.after(100, self._load_dashboard_data)

    def show_cell_analysis(self):
        """Show the cell analysis view"""
        self.analysis_frame.pack_forget()
        self.dashboard_frame.pack_forget()
        self.cell_analysis_frame.pack(fill="both", expand=True)
        self._update_button_states("cell_analysis")

        # Update cell analysis if we have data
        if self.cell_details:
            self._update_cell_analysis()

    def _load_dashboard_data(self):
        """Load dashboard data in background"""
        if self.summary_data:
            # Update dashboard with existing data
            self._update_dashboard()
        else:
            # If no data, show empty state
            for card in self.summary_cards.values():
                card.update_value("N/A")

            self.health_chart.ax.clear()
            self.health_chart.ax.text(0.5, 0.5, "No analysis data",
                                      ha='center', va='center')
            self.health_chart.canvas.draw()

            self.status_chart.ax.clear()
            self.status_chart.ax.text(0.5, 0.5, "No analysis data",
                                      ha='center', va='center')
            self.status_chart.canvas.draw()

    def _update_button_states(self, active_tab):
        """Update navigation button states"""
        self.dashboard_btn.state(["!pressed"])
        self.analysis_btn.state(["!pressed"])
        self.cell_analysis_btn.state(["!pressed"])

        if active_tab == "dashboard":
            self.dashboard_btn.state(["pressed"])
        elif active_tab == "analysis":
            self.analysis_btn.state(["pressed"])
        elif active_tab == "cell_analysis":
            self.cell_analysis_btn.state(["pressed"])

    # ====== Core Functionality ======
    def _browse_file(self):
        """Handle file browsing"""
        filename = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx")],
            title="Select KPI Data File"
        )
        if filename:
            self.file_var.set(filename)
            self.current_file = filename

    def _open_rule_editor(self):
        """Open rule management window"""
        RuleEditor(self)

    def _start_analysis(self):
        """Start analysis in background thread with better feedback"""
        if not self.file_var.get():
            messagebox.showerror("Error", "Please select an input file")
            return

        tech = self.tech_var.get()
        if not tech:
            messagebox.showerror("Error", "Please select a technology")
            return

        # Disable UI during processing
        for child in self.analysis_frame.winfo_children():
            if isinstance(child, ttk.Button):
                child.config(state="disabled")

        # Show loading state
        self.progress_bar.start()
        self.status_label = ttk.Label(
            self.analysis_frame, text="Analyzing data...")
        self.status_label.pack(pady=5)
        self.update_idletasks()

        # Store selected technology
        self.selected_tech = tech

        # Run in background thread
        threading.Thread(
            target=self._run_analysis,
            daemon=True
        ).start()

    def _run_analysis(self):
        """Analyze data and update dashboard"""
        try:
            file_path = self.file_var.get()
            tech = self.selected_tech

            # Clear previous cache for this file
            cache_key = f"{file_path}_{tech}"
            if cache_key in self.analyzer.analysis_cache:
                del self.analyzer.analysis_cache[cache_key]
            if cache_key in self.analyzer.data_cache:
                del self.analyzer.data_cache[cache_key]

            # Perform analysis
            self.summary_data, self.cell_details = self.analyzer.analyze_technology(
                file_path, tech)

            if self.summary_data is None:
                self.status_queue.put(
                    ("error", "Failed to analyze the data file"))
                return

            # Update UI with results
            self.status_queue.put(("update", None))

        except Exception as e:
            self.status_queue.put(("error", f"Analysis failed:\n{str(e)}"))

    def _process_status_queue(self):
        """Process messages from the status queue"""
        try:
            while not self.status_queue.empty():
                msg_type, msg_content = self.status_queue.get_nowait()

                if msg_type == "error":
                    messagebox.showerror("Error", msg_content)
                elif msg_type == "update":
                    # Update analysis complete
                    self.progress_bar.stop()
                    if hasattr(self, 'status_label'):
                        self.status_label.destroy()

                    # Enable UI
                    for child in self.analysis_frame.winfo_children():
                        if isinstance(child, ttk.Button):
                            child.config(state="normal")

                    # Update dashboard if we're viewing it
                    if self.dashboard_frame.winfo_ismapped():
                        self._load_dashboard_data()

        finally:
            self.after(100, self._process_status_queue)

    def _update_dashboard(self):
        """Update dashboard with analysis results"""
        try:
            # Update summary cards
            self.summary_cards["total"].update_value(
                str(self.summary_data["total_cells"]))
            self.summary_cards["healthy"].update_value(
                str(self.summary_data["healthy"]))
            self.summary_cards["warning"].update_value(
                str(self.summary_data["warning"]))
            self.summary_cards["critical"].update_value(
                str(self.summary_data["critical"]))

            # Update health distribution chart
            sizes = [
                self.summary_data["healthy"],
                self.summary_data["warning"],
                self.summary_data["critical"]
            ]
            labels = [
                f"Healthy\n{self.summary_data['healthy']}",
                f"Warning\n{self.summary_data['warning']}",
                f"Critical\n{self.summary_data['critical']}"
            ]
            colors = [
                AppConfig.COLORS['success'],
                AppConfig.COLORS['warning'],
                AppConfig.COLORS['danger']
            ]

            self.health_chart.ax.clear()
            self.health_chart.ax.pie(sizes, labels=labels, colors=colors, autopct="%1.1f%%",
                                     startangle=90, wedgeprops={"edgecolor": "white", "linewidth": 1})
            self.health_chart.ax.set_title("Cell Health Distribution")
            self.health_chart.fig.tight_layout()
            self.health_chart.canvas.draw()

            # Update KPI selector for status chart
            kpis = list(
                set(cell["KPI"] for cell in self.cell_details)) if self.cell_details else []
            self.kpi_chart_combo["values"] = kpis
            if kpis:
                self.kpi_chart_var.set(kpis[0])
                self._update_status_chart()

        except Exception as e:
            print(f"Error updating dashboard: {e}")

    def _update_status_chart(self, event=None):
        """Update the status over time chart based on selected KPI"""
        if not self.summary_data or not self.cell_details:
            return

        kpi = self.kpi_chart_var.get()
        if not kpi:
            return

        # Get the worst performing cells for this KPI
        n_cells = self.cell_count_var.get()
        worst_cells = self.analyzer.get_worst_cells_for_kpi(
            self.cell_details, kpi, n_cells)

        # Prepare data for chart
        days = ["d1", "d2", "d3", "d4", "d5", "d6", "d7"]
        x_labels = [f"Day {i+1}" for i in range(7)]

        chart_data = []
        for cell in worst_cells:
            values = []
            for day in days:
                val = cell.get(day, "No Data")
                if val == "No Data":
                    values.append(np.nan)
                else:
                    try:
                        values.append(float(val))
                    except (ValueError, TypeError):
                        values.append(np.nan)

            chart_data.append({
                'label': cell["Cell Name"],
                'values': values
            })

        # Update chart
        self.status_chart.update_chart(
            chart_data,
            x_labels,
            title=f"Worst {n_cells} Cells for {kpi}",
            ylabel="KPI Value"
        )

    def _update_cell_analysis(self):
        """Update cell analysis view with data"""
        # Clear existing data
        self.cell_tree.delete(*self.cell_tree.get_children())

        # Get unique KPIs for filter
        kpis = list(set(cell["KPI"] for cell in self.cell_details))
        self.kpi_combo["values"] = ["All"] + sorted(kpis)
        self.kpi_var.set("All")
        self.status_var.set("All")

        # Add all cells to treeview
        for cell in self.cell_details:
            self._add_cell_to_tree(cell)

        # Apply initial filters
        self._filter_cells()

    def _add_cell_to_tree(self, cell):
        """Add a cell to the treeview"""
        values = [
            cell["Cell Name"],
            cell["KPI"],
            cell["Status"],
            cell["Score"],
            cell["Bad_days"],
            cell.get("d1", "No Data"),
            cell.get("d1_count", ""),
            cell.get("d2", "No Data"),
            cell.get("d2_count", ""),
            cell.get("d3", "No Data"),
            cell.get("d3_count", ""),
            cell.get("d4", "No Data"),
            cell.get("d4_count", ""),
            cell.get("d5", "No Data"),
            cell.get("d5_count", ""),
            cell.get("d6", "No Data"),
            cell.get("d6_count", ""),
            cell.get("d7", "No Data"),
            cell.get("d7_count", "")
        ]

        item = self.cell_tree.insert("", "end", values=values)

        # Color code based on status
        if cell["Status"] == "Critical":
            self.cell_tree.tag_configure("critical", background="#ffdddd")
            self.cell_tree.item(item, tags=("critical",))
        elif cell["Status"] == "Warning":
            self.cell_tree.tag_configure("warning", background="#fff3cd")
            self.cell_tree.item(item, tags=("warning",))

    def _filter_cells(self):
        """Filter cells based on selected criteria"""
        status_filter = self.status_var.get()
        kpi_filter = self.kpi_var.get()

        # Show all items first
        for item in self.cell_tree.get_children():
            self.cell_tree.item(item, tags=("visible",))

        # Apply status filter
        if status_filter != "All":
            for item in self.cell_tree.get_children():
                values = self.cell_tree.item(item)["values"]
                if values[2] != status_filter:  # Status is at index 2
                    self.cell_tree.item(item, tags=("hidden",))

        # Apply KPI filter
        if kpi_filter != "All":
            visible_items = [item for item in self.cell_tree.get_children()
                             if "visible" in self.cell_tree.item(item)["tags"]]

            for item in visible_items:
                values = self.cell_tree.item(item)["values"]
                if values[1] != kpi_filter:  # KPI is at index 1
                    self.cell_tree.item(item, tags=("hidden",))

        # Hide filtered items
        self.cell_tree.tag_configure("hidden", foreground="gray90")
        self.cell_tree.tag_configure("visible", foreground="black")

    def _export_full_report(self):
        """Export full analysis report to Excel with conditional formatting"""
        if not self.summary_data or not self.cell_details:
            messagebox.showerror("Error", "No analysis results to export")
            return

        output_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save Analysis Report As...",
            initialfile=f"{self.selected_tech}_analysis.xlsx"
        )

        if not output_path:
            return

        try:
            wb = Workbook()

            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"

            # Add summary information
            ws_summary.append(["Technology", self.selected_tech])
            ws_summary.append(
                ["Analysis Date", self.summary_data["timestamp"]])
            ws_summary.append(
                ["Total Cells", self.summary_data["total_cells"]])
            ws_summary.append(["Healthy Cells", self.summary_data["healthy"]])
            ws_summary.append(["Warning Cells", self.summary_data["warning"]])
            ws_summary.append(
                ["Critical Cells", self.summary_data["critical"]])
            ws_summary.append(
                ["Engeneer Overal Comment", ""])
            # Format summary sheet
            for row in ws_summary.iter_rows():
                for cell in row:
                    cell.font = Font(bold=True)

            # Create a sheet for each KPI
            kpis = {cell["KPI"] for cell in self.cell_details}

            # Define formatting styles
            good_fill = PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            bad_fill = PatternFill(start_color="FFC7CE",
                                   end_color="FFC7CE", fill_type="solid")
            warning_fill = PatternFill(
                start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            critical_fill = PatternFill(
                start_color="FF9900", end_color="FF9900", fill_type="solid")
            bold_font = Font(bold=True)

            for kpi in kpis:
                # Filter cells for this KPI
                kpi_cells = [
                    cell for cell in self.cell_details if cell["KPI"] == kpi]

                # Get the rule for this KPI
                rule = next(
                    r for r in self.analyzer.rules[self.selected_tech] if r["kpi"] == kpi)
                op_func = self.analyzer.ops[rule["operator"]]

                # Create worksheet
                ws = wb.create_sheet(title=kpi[:30])  # Limit sheet name length

                # Prepare headers
                headers = ["Cell Name", "Status",
                           "Score", "Bad Days", "Last 5 Days"]
                days = ["d1", "d2", "d3", "d4", "d5", "d6", "d7"]
                for day in days:
                    headers.extend([f"{day.upper()}", f"{day.upper()}_count"])
                headers.append("comment")
                ws.append(headers)

                # Format headers
                for cell in ws[1]:
                    cell.font = bold_font

                # Add data
                for cell in kpi_cells:
                    row = [
                        cell["Cell Name"],
                        cell["Status"],
                        cell["Score"],
                        cell["Bad_days"],
                        cell["Last_5_days"]
                    ]

                    for day in days:
                        row.extend([cell.get(day, "No Data"),
                                    cell.get(f"{day}_count", "")])

                    ws.append(row)

                # Apply conditional formatting
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                    # Format status column
                    status_cell = ws.cell(row=row_idx, column=2)
                    if status_cell.value == "Critical":
                        status_cell.fill = critical_fill
                    elif status_cell.value == "Warning":
                        status_cell.fill = warning_fill

                    # Format KPI values
                    # Only the value columns (skip counts)
                    for col_idx in range(6, 20, 2):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        try:
                            value = float(cell.value)
                            if op_func(value, rule["threshold"]):
                                cell.fill = good_fill
                            else:
                                cell.fill = bad_fill
                        except (ValueError, TypeError):
                            pass

            wb.save(output_path)
            messagebox.showinfo(
                "Success", f"Analysis exported successfully to:\n{output_path}")

        except Exception as e:
            messagebox.showerror(
                "Error", f"Failed to export results:\n{str(e)}")

    def _export_selected_cells(self):
        """Export selected cells to Excel"""
        selected_items = self.cell_tree.selection()
        if not selected_items:
            messagebox.showerror("Error", "Please select cells to export")
            return

        output_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save Selected Cells As...",
            initialfile=f"{self.selected_tech}_selected_cells.xlsx"
        )

        if not output_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Selected Cells"

            # Prepare headers
            headers = ["Cell Name", "KPI", "Status", "Score", "Bad Days"]
            days = ["d1", "d2", "d3", "d4", "d5", "d6", "d7"]
            for day in days:
                headers.extend([f"{day.upper()}", f"{day.upper()}_count"])

            ws.append(headers)

            # Format headers
            for cell in ws[1]:
                cell.font = Font(bold=True)

            # Define formatting styles
            warning_fill = PatternFill(
                start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            critical_fill = PatternFill(
                start_color="FF9900", end_color="FF9900", fill_type="solid")

            # Add data
            for item in selected_items:
                values = self.cell_tree.item(item)["values"]
                ws.append(values)

                # Apply formatting based on status
                status = values[2]  # Status is at index 2
                for cell in ws[ws.max_row]:
                    if status == "Critical":
                        cell.fill = critical_fill
                    elif status == "Warning":
                        cell.fill = warning_fill

            wb.save(output_path)
            messagebox.showinfo(
                "Success", f"Selected cells exported to:\n{output_path}")

        except Exception as e:
            messagebox.showerror(
                "Error", f"Failed to export selected cells:\n{str(e)}")

# ====== Rule Editor ======


class RuleEditor(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("KPI Rule Editor")
        self.geometry("1000x750")
        self.configure(bg=BG_COLOR)
        self.current_edit_index = None
        self.current_tech = None

        # Header
        header_frame = tk.Frame(self, bg=PRIMARY_COLOR)
        header_frame.pack(fill="x", padx=10, pady=(10, 0))
        tk.Label(header_frame, text="KPI Rule Configuration",
                 font=(FONT_NAME, 14, "bold"), bg=PRIMARY_COLOR, fg="white").pack(pady=10)

        # Notebook for technologies
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # Create tabs for each technology
        self.tech_tabs = {}
        for tech in ["2G", "3G", "4G"]:
            tab = ttk.Frame(self.notebook)
            self.notebook.add(tab, text=tech)
            self.tech_tabs[tech] = tab
            self._create_tech_tab(tab, tech)

        # Save button
        save_frame = ttk.Frame(self)
        save_frame.pack(fill="x", padx=10, pady=(0, 10))
        ModernButton(save_frame, text="Save Changes",
                     command=self._save_and_close).pack()

        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_change)

    def _on_tab_change(self, event):
        """Handle tab change event"""
        tab_id = self.notebook.select()
        tab_text = self.notebook.tab(tab_id, "text")
        self.current_tech = tab_text

    def _create_tech_tab(self, tab, tech):
        """Create a tab for each technology with enhanced fields"""
        # Input frame
        input_frame = ttk.Frame(tab, padding=10)
        input_frame.pack(fill="x", pady=5)

        # KPI Name
        ttk.Label(input_frame, text="KPI Name:").grid(
            row=0, column=0, padx=(0, 5), sticky="e")
        kpi_entry = ttk.Entry(input_frame, width=35)
        kpi_entry.grid(row=0, column=1, padx=5, sticky="ew")

        # Operator
        ttk.Label(input_frame, text="Operator:").grid(
            row=0, column=2, padx=(10, 5), sticky="e")
        operator_combo = ttk.Combobox(
            input_frame, values=[">", "<", ">=", "<=", "=="], width=3)
        operator_combo.grid(row=0, column=3, padx=5)
        operator_combo.current(0)

        # Threshold
        ttk.Label(input_frame, text="Threshold:").grid(
            row=0, column=4, padx=(10, 5), sticky="e")
        threshold_entry = ttk.Entry(input_frame, width=8)
        threshold_entry.grid(row=0, column=5, padx=5)

        # Count Column
        ttk.Label(input_frame, text="Count Column:").grid(
            row=1, column=0, padx=(0, 5), sticky="e")
        count_entry = ttk.Entry(input_frame, width=20)
        count_entry.grid(row=1, column=1, padx=5, sticky="ew")

        # Count Threshold
        ttk.Label(input_frame, text="Count Threshold:").grid(
            row=1, column=2, padx=(10, 5), sticky="e")
        count_threshold_entry = ttk.Entry(input_frame, width=8)
        count_threshold_entry.grid(row=1, column=3, padx=5)
        count_threshold_entry.insert(0, "0")

        ttk.Label(input_frame, text="(Leave empty if not needed)").grid(
            row=1, column=4, columnspan=2, sticky="w")

        # Treeview with scrollbars
        tree_frame = ttk.Frame(tab)
        tree_frame.pack(fill="both", expand=True, padx=5, pady=5)

        tree = ttk.Treeview(
            tree_frame,
            columns=("kpi", "operator", "threshold",
                     "count_column", "count_threshold"),
            show="headings",
            selectmode="browse",
            height=20
        )
        tree.heading("kpi", text="KPI Name")
        tree.heading("operator", text="Operator")
        tree.heading("threshold", text="Threshold")
        tree.heading("count_column", text="Count Column")
        tree.heading("count_threshold", text="Count Threshold")
        tree.column("kpi", width=340)
        tree.column("operator", width=60, anchor="center")
        tree.column("threshold", width=80, anchor="center")
        tree.column("count_column", width=340, anchor="center")
        tree.column("count_threshold", width=120, anchor="center")

        yscroll = ttk.Scrollbar(
            tree_frame, orient="vertical", command=tree.yview)
        xscroll = ttk.Scrollbar(
            tree_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscroll=yscroll.set, xscroll=xscroll.set)

        tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")

        # Load existing rules
        for rule in rules_map[tech]:
            tree.insert("", "end", values=(
                rule["kpi"],
                rule["operator"],
                rule["threshold"],
                rule.get("count_column", ""),
                rule.get("count_threshold", "")
            ))

        # Button frame
        btn_frame = ttk.Frame(tab)
        btn_frame.pack(fill="x", padx=5, pady=(0, 10))

        def add_rule():
            """Add new rule validation"""
            kpi = kpi_entry.get().strip()
            operator = operator_combo.get()
            threshold = threshold_entry.get()
            count_col = count_entry.get().strip()
            count_thresh = count_threshold_entry.get()

            if not kpi:
                messagebox.showerror("Error", "Please enter a KPI name")
                return

            if not threshold:
                messagebox.showerror("Error", "Please enter a threshold value")
                return

            try:
                threshold_val = float(threshold)
                count_thresh_val = float(count_thresh) if count_thresh else 0

                rule = {
                    "kpi": kpi,
                    "operator": operator,
                    "threshold": threshold_val,
                    "count_threshold": count_thresh_val
                }

                if count_col:
                    rule["count_column"] = count_col

                if self.current_edit_index is not None:
                    # Update existing rule
                    rules_map[tech][self.current_edit_index] = rule
                    tree.item(tree.selection()[0], values=(
                        kpi, operator, threshold_val,
                        count_col if count_col else "",
                        count_thresh_val
                    ))
                    self.current_edit_index = None
                else:
                    # Add new rule
                    tree.insert("", "end", values=(
                        kpi, operator, threshold_val,
                        count_col if count_col else "",
                        count_thresh_val
                    ))
                    rules_map[tech].append(rule)

                clear_fields()
            except ValueError:
                messagebox.showerror("Error", "Invalid numeric value")

        def edit_rule():
            """Edit selected rule"""
            selected = tree.selection()
            if not selected:
                messagebox.showerror("Error", "Please select a rule to edit")
                return

            index = tree.index(selected[0])
            self.current_edit_index = index
            self.current_tech = tech

            values = tree.item(selected[0])["values"]

            kpi_entry.delete(0, tk.END)
            kpi_entry.insert(0, values[0])

            operator_combo.set(values[1])

            threshold_entry.delete(0, tk.END)
            threshold_entry.insert(0, values[2])

            count_entry.delete(0, tk.END)
            count_entry.insert(0, values[3] if values[3] else "")

            count_threshold_entry.delete(0, tk.END)
            count_threshold_entry.insert(0, values[4] if values[4] else "0")

        def delete_rule():
            """Delete selected rule"""
            selected = tree.selection()
            if not selected:
                messagebox.showerror("Error", "Please select a rule to delete")
                return
            index = tree.index(selected[0])
            del rules_map[tech][index]
            tree.delete(selected[0])
            clear_fields()
            self.current_edit_index = None

        def clear_fields():
            """Clear all input fields"""
            kpi_entry.delete(0, tk.END)
            operator_combo.current(0)
            threshold_entry.delete(0, tk.END)
            count_entry.delete(0, tk.END)
            count_threshold_entry.delete(0, tk.END)
            count_threshold_entry.insert(0, "0")
            self.current_edit_index = None

        # Buttons
        ModernButton(btn_frame, text="Add/Update Rule",
                     command=add_rule).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Edit Selected",
                   command=edit_rule).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Delete Selected",
                   command=delete_rule).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Clear Fields",
                   command=clear_fields).pack(side="right", padx=5)

    def _save_and_close(self):
        """Save rules and close window"""
        save_rules()
        self.destroy()
        messagebox.showinfo("Success", "Rules saved successfully")


def save_rules():
    """Save rules to JSON file"""
    with open(RULES_FILE, 'w') as f:
        json.dump(rules_map, f, indent=4)


# ====== Main Application ======
if __name__ == "__main__":
    app = CellPerformanceApp()
    app.mainloop()
